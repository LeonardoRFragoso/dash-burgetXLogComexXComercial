from openpyxl import load_workbook, Workbook

# Caminho do arquivo original
caminho_arquivo = r"C:\Users\leonardo.fragoso\Desktop\Projetos\dash-burgetXLogComexXComercial\iTRACKER_novo 01.06 v2.xlsx"

# Nome do novo arquivo de saída
novo_arquivo = r"C:\Users\leonardo.fragoso\Desktop\Projetos\dash-burgetXLogComexXComercial\Planilha1_extraida.xlsx"

# Abrir a planilha original no modo somente leitura
wb_original = load_workbook(caminho_arquivo, read_only=True)
ws_original = wb_original["Planilha1"]

# Criar um novo workbook e aba
wb_novo = Workbook()
ws_novo = wb_novo.active
ws_novo.title = "Planilha1"

# Copiar linha por linha da aba original para a nova
for row in ws_original.iter_rows(values_only=True):
    ws_novo.append(row)

# Salvar o novo arquivo
wb_novo.save(novo_arquivo)

print(f"Nova planilha salva com sucesso em: {novo_arquivo}")
